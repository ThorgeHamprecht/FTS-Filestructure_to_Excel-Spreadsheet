import os
import random
from natsort import os_sorted

import tkinter as tk
from threading import Thread
from tkinter import filedialog

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import PatternFill

firstRow = 4
path = ""
working = False

workbook = Workbook()
sheet = workbook.active

# styles
bold_font = Font(bold=True)
cursive_font = Font(italic=True)
bold_cursive_font = Font(bold=True, italic=True)
center_aligned_text = Alignment(horizontal="center")
thin = Side(border_style="thin")
blank = Side(border_style="thin", color="FFFFFF")
border_blank = Border(top=blank, left=blank, right=blank, bottom=blank)

# formatting rules for folders
blue_background = PatternFill(fgColor="DDEBF7")


def walk_internal(path_internal, layer, row):
    layer.append(0)

    with os.scandir(path_internal) as it:
        entries = list(it)
        entries = os_sorted(entries, key=lambda x: x.name)
    for entry in entries:
        if entry.name == ".DS_Store":
            continue

        layer.append(layer.pop(len(layer) - 1) + 1)

        nr = ""
        for num in layer:
            nr += str(num) + "."

        if entry.is_file():
            sheet["B" + str(row[0])] = nr
            sheet["C" + str(row[0])] = entry.name
            sheet["D" + str(row[0])] = "document"

            for rows in sheet.iter_rows(min_row=row[0], max_row=row[0], min_col=1, max_col=1):
                for cell in rows:
                    cell.border = border_blank

            for rows in sheet.iter_rows(min_row=row[0], max_row=row[0], min_col=2, max_col=6):
                for cell in rows:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            for rows in sheet.iter_rows(min_row=row[0], max_row=row[0], min_col=7, max_col=7):
                for cell in rows:
                    cell.border = border_blank

            row[0] += 1
        else:
            sheet["B" + str(row[0])] = nr
            sheet["C" + str(row[0])] = entry.name
            sheet["D" + str(row[0])] = "folder"

            for rows in sheet.iter_rows(min_row=row[0], max_row=row[0], min_col=1, max_col=1):
                for cell in rows:
                    cell.border = border_blank

            for rows in sheet.iter_rows(min_row=row[0], max_row=row[0], min_col=2, max_col=6):
                for cell in rows:
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            for rows in sheet.iter_rows(min_row=row[0], max_row=row[0], min_col=7, max_col=7):
                for cell in rows:
                    cell.border = border_blank

            row[0] += 1

            walk_internal(path_internal + "/" + entry.name, layer, row)
    layer.pop(len(layer) - 1)


def walk():
    global path
    global curRow
    internal_path = path
    walk_internal(internal_path, [], curRow)


# path from which the directory is walked to the bottom
initialPath = ""
savePath = ""

curRow = [firstRow + 1]

bg_color = "#1e1f22"
bt_color = "#366ace"
tx_color = "#a9b7c6"
bt_color_a = "white"
bt_color_tx = "white"
# GUI
window = tk.Tk()

# window info
window.title("FTS - Files To Spreadsheet")
window.configure(bg=bg_color)
# window.iconbitmap("FTS - Logo.ico")
window_width = 600
window_height = 400
window.resizable(False, False)


# get the screen dimension
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()

# find the center point
center_x = int(screen_width / 2 - window_width / 2)
center_y = int(screen_height / 2 - window_height / 2)

# set the position of the window to the center of the screen
window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

label = tk.Label(
    text="Start the generation of your Spreadsheet:",
    foreground="#6a8759",  # Set the text color to white
    background=bg_color,  # Set the background color to black
    font=('Arial', 13)
)
label.pack()

instruction1 = tk.Label(
    text="1.    Choose a directory: ",
    foreground=tx_color,  # Set the text color to white
    background=bg_color,  # Set the background color to black
    font=('Arial', 13)
)
instruction1.pack(anchor="w")

current_dir = tk.Label(
    text="Current path: " + initialPath,
    foreground="#6897bb",  # Set the text color to white
    background=bg_color,  # Set the background color to black
    font=('Arial', 10)
)
current_dir.pack(anchor="w", padx=10)


def browse():
    global working

    if not working:
        global initialPath
        global path
        initialPath = filedialog.askdirectory()
        path = initialPath
        current_dir.config(text="Current path: " + initialPath)


tk.Button(text="Change Directory",
          width=16,
          foreground=bt_color_tx,
          background=bt_color,
          activebackground=bt_color,
          activeforeground=bt_color_a,
          font=('Arial', 13),
          command=browse
          ).pack(anchor="w", padx=15)

instruction2 = tk.Label(
    text="2.    Choose a export location: ",
    foreground=tx_color,  # Set the text color to white
    background=bg_color,  # Set the background color to black
    font=('Arial', 13)
)
instruction2.pack(anchor="w")

current_export = tk.Label(
    text="Current export path: " + savePath,
    foreground="#6897bb",  # Set the text color to white
    background=bg_color,  # Set the background color to black
    font=('Arial', 10)
)
current_export.pack(anchor="w", padx=10)


def browse_path():
    global working
    global savePath
    if not working:
        savePath = filedialog.askdirectory()
        current_export.config(text="Current export path: " + savePath)


tk.Button(text="Change Export Path",
          width=16,
          foreground=bt_color_tx,
          background=bt_color,
          activebackground=bt_color,
          activeforeground=bt_color_a,
          font=('Arial', 13),
          command=browse_path
          ).pack(anchor="w", padx=15)


def create_spreadsheet():
    global initialPath
    global savePath
    global messageUser

    if initialPath != "" and savePath != "":
        global working
        global window
        global workbook
        global sheet
        global curRow

        workbook = Workbook()
        sheet = workbook.active
        curRow = [firstRow + 1]

        columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
        dimensions = [4, 15, 85, 10, 11, 25, 1000, 4, 4]

        for index in range(0, len(columns)):
            sheet.column_dimensions[columns[index]].width = dimensions[index]
            if index == 0:
                for rows in sheet.iter_rows(min_row=1, max_row=firstRow - 1, min_col=1, max_col=7):
                    for cell in rows:
                        cell.border = border_blank

        sheet["B" + str(firstRow)] = "No."
        sheet["B" + str(firstRow)].font = bold_cursive_font
        sheet["B" + str(firstRow)].alignment = center_aligned_text

        sheet["C" + str(firstRow)] = "Document Name"
        sheet["C" + str(firstRow)].font = bold_cursive_font
        sheet["C" + str(firstRow)].alignment = center_aligned_text

        sheet["D" + str(firstRow)] = "Type"
        sheet["D" + str(firstRow)].font = bold_cursive_font
        sheet["D" + str(firstRow)].alignment = center_aligned_text

        sheet["E" + str(firstRow)] = "reviewed"
        sheet["E" + str(firstRow)].font = bold_cursive_font
        sheet["E" + str(firstRow)].alignment = center_aligned_text

        sheet["F" + str(firstRow)] = "remark"
        sheet["F" + str(firstRow)].font = bold_cursive_font
        sheet["F" + str(firstRow)].alignment = center_aligned_text

        for rows in sheet.iter_rows(min_row=firstRow, max_row=firstRow, min_col=2, max_col=6):
            for cell in rows:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        working = True

        thread = Thread(target=walk())
        thread.start()

        messageUser.config(text="Creating Spreadsheet...", foreground=tx_color)

        thread.join()

        for rows in sheet.iter_rows(min_row=curRow[0], max_row=curRow[0], min_col=1, max_col=7):
            for cell in rows:
                cell.border = border_blank
        sheet.row_dimensions[curRow[0]].height = 100000000

        doc_id = str(random.randint(0, 1000))
        workbook.save(savePath + "/Spreadsheet" + doc_id + ".xlsx")
        working = False
        messageUser.config(text="Created Spreadsheet!", foreground="#7676ac")

    else:
        messageUser.config(text="MUST FIRST ASSIGN A DIRECTORY AND SAVE LOCATION!", foreground="#eb7171")


tk.Button(text="Create Spreadsheet",
          width=16,
          foreground=bt_color_tx,
          background="#57965c",
          activebackground="#aa4926",
          activeforeground=bt_color_a,
          font=('Arial', 13),
          command=create_spreadsheet
          ).pack(anchor="w", pady=20, padx=15)

messageUser = tk.Label(
            text="",
            foreground=tx_color,  # Set the text color to white
            background=bg_color,  # Set the background color to black
            font=('Arial', 13)
)
messageUser.pack(anchor="w", padx=10)

window.mainloop()
